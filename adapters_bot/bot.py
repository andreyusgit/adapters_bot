import logging

from openpyxl import Workbook, load_workbook

from aiogram import Bot, types
from aiogram.utils import executor
from aiogram.dispatcher import Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.contrib.middlewares.logging import LoggingMiddleware

from config import TOKEN
from utils import TestStates
from messages import MESSAGES

logging.basicConfig(format=u'%(filename)+13s [ LINE:%(lineno)-4s] %(levelname)-8s [%(asctime)s] %(message)s',
                    level=logging.DEBUG)

bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
users = {}
univer = ['ФИЗМЕХ', 'ГИ', 'ИБСИБ', 'ИЭ', 'ИЭИТ', 'ИКИЗИ', 'ИКНТ', 'ИММИТ', 'ИПМЭИТ', 'ИСИ', 'ИСПО']

dp.middleware.setup(LoggingMiddleware())


@dp.message_handler(commands=['start'])
async def process_start_command(message: types.Message):
    await message.answer(MESSAGES['start'])


@dp.message_handler(state='*', commands=['help'])
async def process_help_command(message: types.Message):
    await message.answer(MESSAGES['help'])


@dp.message_handler(state='*', commands=['thanks'])
async def process_help_command(message: types.Message):
    await message.answer(MESSAGES['thx'])


@dp.message_handler(state='*', commands=['password'])
async def process_setstate_command(message: types.Message):
    uni = -1
    argument = message.get_args()
    state = dp.current_state(user=message.from_user.id)
    if not argument:
        await state.reset_state()
        return await message.answer(MESSAGES['state_reset'])

    if not argument.isdigit():
        return await message.answer(MESSAGES['invalid_key'].format(key=argument))
    elif argument == '271':
        uni = 0
        await state.set_state(TestStates.all()[1])
    elif argument == '092':
        uni = 1
        await state.set_state(TestStates.all()[1])
    elif argument == '192':
        uni = 2
        await state.set_state(TestStates.all()[1])
    elif argument == '176':
        uni = 3
        await state.set_state(TestStates.all()[1])
    elif argument == '099':
        uni = 4
        await state.set_state(TestStates.all()[1])
    elif argument == '921':
        uni = 5
        await state.set_state(TestStates.all()[1])
    elif argument == '736':
        uni = 6
        await state.set_state(TestStates.all()[1])
    elif argument == '826':
        uni = 7
        await state.set_state(TestStates.all()[1])
    elif argument == '645':
        uni = 8
        await state.set_state(TestStates.all()[1])
    elif argument == '725':
        uni = 9
        await state.set_state(TestStates.all()[1])
    elif argument == '062':
        uni = 10
        await state.set_state(TestStates.all()[1])
    else:
        return await message.answer(MESSAGES['invalid_key'].format(key=argument))

    users[int(message.from_user.id)] = uni
    await message.answer(MESSAGES['state_change'].format(key=univer[uni]))


@dp.message_handler(state=TestStates.TEST_STATE_1)
async def first_test_state_case_met(message: types.Message):
    wbSearch = Workbook()
    max_str = 0
    uni = users[int(message.from_user.id)]
    if uni == 0:
        wbSearch = load_workbook("FHIS.xlsx")
        max_str = 95
    elif uni == 1:
        wbSearch = load_workbook("GI.xlsx")
        max_str = 121
    elif uni == 2:
        wbSearch = load_workbook("IBSIB.xlsx")
        max_str = 46
    elif uni == 3:
        wbSearch = load_workbook("IE.xlsx")
        max_str = 90
    elif uni == 4:
        wbSearch = load_workbook("IEIT.xlsx")
        max_str = 70
    elif uni == 5:
        wbSearch = load_workbook("IKIZI.xlsx")
        max_str = 40
    elif uni == 6:
        wbSearch = load_workbook("IKNT.xlsx")
        max_str = 126
    elif uni == 7:
        wbSearch = load_workbook("IMMIT.xlsx")
        max_str = 72
    elif uni == 8:
        wbSearch = load_workbook("IPMEIT.xlsx")
        max_str = 196
    elif uni == 9:
        wbSearch = load_workbook("ISI.xlsx")
        max_str = 152
    elif uni == 10:
        wbSearch = load_workbook("ISPO.xlsx")
        max_str = 158

    wsSearch = wbSearch.active

    index = 0
    for i in range(2, max_str):
        check = str(message.text).lower().split()
        value = wsSearch.cell(row=i, column=1).value
        value = str(value).split()
        if (value.pop(0) == check.pop(0)) and (value.pop(0) == check.pop(0)):
            index = i
            break
        if i == (max_str - 1):
            an = 'Поиски не принесли результата\nПроверьте правильность ФИО или попробуйте написать имя в сокращении'
            await message.answer(an)
            return
    if index != 0:
        text = str(wsSearch.cell(row=index, column=1).value) + '\n' + str(wsSearch.cell(row=index, column=2).value)
        if (uni != 2) and (uni != 4) and (uni != 8):
            text = text + '\n' + str(wsSearch.cell(row=index, column=4).value) + '\n'
        text = text + str(wsSearch.cell(row=index, column=5).value) + '\n'
        text = text + str(wsSearch.cell(row=index, column=6).value) + '\n'
        text = text + str(wsSearch.cell(row=index, column=7).value) + '\n'
        text = text + '\n*Мотивационное письмо:*\n\n'
        text = text + str(wsSearch.cell(row=index, column=8).value) + '\n'
        if (uni == 0) or (uni == 2) or (uni == 3) or (uni == 4) or (uni == 5) or (uni == 6) or (uni == 8) or (uni == 9):
            text = text + '\n' + str(wsSearch.cell(row=index, column=10).value) + '\n'
        elif uni == 1:
            text = text + '\n*Сильные и слабые стороны:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=9).value) + '\n'
            text = text + '\n*Что в работе твоих адаптеров понравилось больше всего:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=10).value) + '\n'
            text = text + '\n*Где состоишь/состоял:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=11).value) + '\n\n'
            text = text + str(wsSearch.cell(row=index, column=13).value) + '\n'
        elif uni == 7:
            text = text + '\n*Как расставлены твои приоритеты:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=9).value) + '\n\n'
            text = text + str(wsSearch.cell(row=index, column=11).value) + '\n'
        else:
            text = text + '\n*Краткий автопортрет:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=9).value) + '\n'
            text = text + '\n*Кто ты из смешариков:*\n\n'
            text = text + str(wsSearch.cell(row=index, column=10).value) + '\n\n'
            text = text + str(wsSearch.cell(row=index, column=12).value) + '\n'
        await message.answer(text)
        tt = '/thanks - сказать спасибо\nЧтобы найти нового человека - введите ФИО еще раз'
        await message.answer(tt)


@dp.callback_query_handler(text="random_value")
async def send_random_value(call: types.CallbackQuery):
    await call.answer(text="Спасибо, что воспользовались ботом!", show_alert=True)


@dp.message_handler(state=TestStates.all())
async def some_test_state_case_met(message: types.Message):
    await message.answer('К сожаленю, я не пока не зеаю такую команду')

    


@dp.message_handler()
async def echo_message(msg: types.Message):
    await bot.send_message(msg.from_user.id, 'К сожаленю, я не могу работать пока вы не введете пароль')


async def shutdown(dispatcher: Dispatcher):
    await dispatcher.storage.close()
    await dispatcher.storage.wait_closed()


if __name__ == '__main__':
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button_1 = types.KeyboardButton(text="/start")
    keyboard.add(button_1)
    executor.start_polling(dp, on_shutdown=shutdown)
